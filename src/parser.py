from time import time
from openpyxl import load_workbook
from string import ascii_uppercase

from src.database import DataBase

class XLSXParser:

    COLUMNS = ['', *ascii_uppercase]

    def __init__(self, filename):
        self.workbook = load_workbook(filename)
        self.worksheets = self.workbook.worksheets

    def parse_categories_names(self):
        return [sheet.title for sheet in self.worksheets]

    def parse_addvanced_attributes_names(self):
        attributes = set()
        for sheet in self.worksheets:
            attributes |= set(self.get_adv_attrs_by_category(sheet.title))
        return list(attributes)

    def parse_brands_names(self): 
        brands = set() 
        for sheet in self.worksheets: 
            sheet_cells = sheet['C2': 'C' + str(sheet.max_row)] 
            brands |= set([row_cells[0].value for row_cells in sheet_cells]) 
        return list(brands)

    def generator_parse_proructs_row(self, category):
        sheet = self.workbook[category]
        end_cell = self.COLUMNS[sheet.max_column] + str(sheet.max_row)
        sheet_data_cells = sheet['B3': end_cell]

        for row_cells in sheet_data_cells:
            yield [cell.value for cell in row_cells]

    def get_adv_attrs_by_category(self, category):
        sheet = self.workbook[category]
        from_column = self.COLUMNS[6] + '2'
        to_column = self.COLUMNS[sheet.max_column] + '2'
        cells = sheet[from_column:to_column][0]
        return [cell.value for cell in cells]





