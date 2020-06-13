import openpyxl
from openpyxl.styles import Font

from .tools import ExcellGenTools



class GeneratorXLSX:
    title_font= Font(name='Arial', bold=True)

    def __call__(self, cli_args):

        self._init_workbook()

        for index, category in enumerate(ExcellGenTools.CATEGORIES):
            self._create_category_sheet(category, index)
            self._fill_data(category, cli_args.rows)

        self.workbook.save(cli_args.file)

    def _init_workbook(self):
        self.workbook = openpyxl.Workbook()
        _sheet = self.workbook.active
        self.workbook.remove(_sheet)

    def _create_category_sheet(self, category, index):
        self.workbook.create_sheet(title=category, index=index)
        sheet = self.workbook[category]
        for column, product_attr in enumerate(ExcellGenTools.PRODUCTS, start=2):
            cell = sheet.cell(row=2, column=column)
            cell.value = product_attr
            cell.font = self.title_font
            column_width = ExcellGenTools.get_column_width(product_attr)
            sheet.column_dimensions[ExcellGenTools.COLUMNS[column]].width = column_width

        for column, adv_product_attr in enumerate(
            ExcellGenTools.get_attribute_list(category), start=len(ExcellGenTools.CATEGORIES) + 2
        ):
            cell = sheet.cell(row=2, column=column)
            cell.value = adv_product_attr
            cell.font = self.title_font
            column_width = ExcellGenTools.get_column_width(adv_product_attr)
            if column_width:
                sheet.column_dimensions[ExcellGenTools.COLUMNS[column]].width = column_width

    def _fill_data(self, category, rows):
        sheet = self.workbook[category]
        for row in range(3, rows + 3):
            brand = ExcellGenTools.get_brand()
            sheet.cell(row=row, column=2).value = ExcellGenTools.get_article()
            sheet.cell(row=row, column=3).value = brand
            sheet.cell(row=row, column=4).value = ExcellGenTools.get_name(category, brand)
            sheet.cell(row=row, column=5).value = ExcellGenTools.get_price()
            for column, adv_attr in enumerate(
                ExcellGenTools.get_attribute_list(category), start=len(ExcellGenTools.CATEGORIES) + 2
            ):
                sheet.cell(row=row, column=column).value = ExcellGenTools.get_attribute_value(adv_attr)
