import openpyxl
from openpyxl.styles import Font

from .tools import ExcellGenTools



class GeneratorXLSX:
    title_font= Font(name='Arial', bold=True)

    def __call__(self, cli_args):

        self._init_workbook()

        for index, category in enumerate(ExcellGenTools.CATEGORIES):
            self._create_category_sheet(category, index)

        self.workbook.save(cli_args.file)

    def _init_workbook(self):
        self.workbook = openpyxl.Workbook()
        _sheet = self.workbook.active
        self.workbook.remove(_sheet)

    def _create_category_sheet(self, category, index):
        self.workbook.create_sheet(title=category, index=index)
        sheet = self.workbook[category]
        for column, product_atr in enumerate(ExcellGenTools.PRODUCTS, start=2):
            cell = sheet.cell(row=2, column=column)
            cell.value = product_atr
            cell.font = self.title_font
        sheet.column_dimensions['D'].width = 20

        for column, adv_product_atr in enumerate(
            ExcellGenTools.CATEGORIES[category][ExcellGenTools.ATRIBUT_LIST],
            start=2 + len(ExcellGenTools.CATEGORIES)
        ):
            cell = sheet.cell(row=2, column=column)
            cell.value = adv_product_atr
            cell.font = self.title_font
            column_width = ExcellGenTools.ATTRS[adv_product_atr][ExcellGenTools.COLUMN_WIDTH]
            if column_width:
                sheet.column_dimensions[ExcellGenTools.COLUMNS[column]].width = column_width
