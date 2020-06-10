import asyncio
from time import time

import aiojobs
import openpyxl
import uvloop

from database.engine import DataBase, AsyncPostgreSQL

CHUNK_READ_RAWS = 100
MAX_DB_CONNECTION = 20

class AsyncXLSXParser:

    PRODUCT_FIELDS = ['category_id', 'brand_id', 'product_code', 'name', 'price']
    ATTRIBUTE_FIELDS = ['product_code', 'attribute_id', 'attribute_value']
    TABLENAME_CATEGORIES = 'categories'
    TABLENAME_ATTRIBUTES = 'attributes'
    TABLENAME_BRANDS = 'brands'
    TABLENAME_PRODUCTS = 'products'
    TABLENAME_PRODUCT_ATTRIBUTES = 'product_attributes'

    def __init__(self, loop, cli_args):
        self.cli_args = cli_args
        self.loop = loop
        self.workbook = openpyxl.load_workbook(self.cli_args.file)
        self.worksheets = self.workbook.worksheets
        self.categories = [ws.title for ws in self.worksheets]
        self.brand_lock = asyncio.Lock(loop=loop)
        self.attribute_lock = asyncio.Lock(loop=loop)
        self.brand_ids = {}
        self.product_attribute_ids = {}

    async def run(self):
        self.scheduler = await aiojobs.create_scheduler(limit=MAX_DB_CONNECTION)
        await asyncio.gather(
            *[self.parse_category(sheet) for sheet in self.worksheets]
        )
        pending = asyncio.Task.all_tasks()
        asyncio.gather(*pending)

    async def parse_category(self, sheet):
        database = AsyncPostgreSQL()
        async with database:
            category_id = await database.create(
                tablename=self.TABLENAME_CATEGORIES,
                fields=['category_name'],
                values=[sheet.title],
                return_value='category_id',
            )

        category_attrs_list = []
        for column in range(6, sheet.max_column + 1):
            attribute = sheet.cell(row=2, column=column).value
            if attribute not in self.product_attribute_ids:
                async with self.attribute_lock:
                    await self._check_attribute(attribute)
            category_attrs_list.append(self.product_attribute_ids[attribute])

        read_from_row = 3
        read_to_row = read_from_row
        while read_to_row <= sheet.max_row:
            read_to_row += CHUNK_READ_RAWS
            if read_to_row > sheet.max_row:
                read_to_row = sheet.max_row + 1

            products = []
            advanced_attributes = []
            for row in range(read_from_row, read_to_row):
                brand = sheet.cell(row=row, column=3).value
                if brand not in self.brand_ids:
                    async with self.brand_lock:
                        await self._check_brand(brand)
                products.append(
                    (
                        category_id,
                        self.brand_ids[brand],
                        sheet.cell(row=row, column=2).value,
                        sheet.cell(row=row, column=4).value,
                        sheet.cell(row=row, column=5).value,
                    )
                )
                attr_values = [sheet.cell(row=row, column=i).value for i in range(6, sheet.max_column + 1)]
                for attribute_id, attribute_value in zip(category_attrs_list, attr_values):
                    advanced_attributes.append(
                        (
                            sheet.cell(row=row, column=2).value,
                            attribute_id,
                            attribute_value,
                        )
                    )
            await self.scheduler.spawn(self._create_products_and_attrs(products, advanced_attributes))

            read_from_row = read_to_row


    async def _check_attribute(self, attribute):
        if attribute not in self.product_attribute_ids:
            db = AsyncPostgreSQL()
            async with db:
                self.product_attribute_ids[attribute] = await db.create(
                    tablename=self.TABLENAME_ATTRIBUTES,
                    fields=['attribute_name'],
                    values=[attribute],
                    return_value='attribute_id',
                )

    async def _check_brand(self, brand):
        if brand not in self.brand_ids:
            db = AsyncPostgreSQL()
            async with db:
                self.brand_ids[brand] =  await db.create(
                    tablename=self.TABLENAME_BRANDS,
                    fields=['brand_name'],
                    values=[brand],
                    return_value='brand_id',
                )

    async def _create_products_and_attrs(self, products, advanced_attributes):
        db = AsyncPostgreSQL()
        async with db:
            await db.bulk_create(self.TABLENAME_PRODUCTS, self.PRODUCT_FIELDS, products)
            await db.bulk_create(
                self.TABLENAME_PRODUCT_ATTRIBUTES, self.ATTRIBUTE_FIELDS, advanced_attributes
            )

def run(cli_args):
    uvloop.install()
    loop = asyncio.get_event_loop()
    parser = AsyncXLSXParser(loop, cli_args)
    start = time()
    loop.run_until_complete(parser.run())
    print(f'ASYNC PARSE DONE on {time()-start:.3f}')