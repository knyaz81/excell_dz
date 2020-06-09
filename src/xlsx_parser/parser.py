from time import time
import openpyxl

from database.engine import DataBase

class XLSXParser:

    def __call__(self, cli_args):
        self.cli_args = cli_args
        self.workbook = openpyxl.load_workbook(self.cli_args.file)
        self.worksheets = self.workbook.worksheets
        self.categories = [ws.title for ws in self.worksheets]
        self.brand_ids = {}
        self.prod_attr_ids = {}
        self.database = DataBase()

        self.database.init_db()
        self.database.create_tables()

        overall_start = time()
        for sheet in self.worksheets:
            start = time()
            self._parser_category_sheet(sheet)
            print(f'PARSE category "{sheet.title} DONE in {(time()-start):.3f}sec"\n')
        print(f'DOCUMENT PARSE DONE in {(time()-overall_start):3f}sec')

    def _parser_category_sheet(self, sheet):
        category_id = self.database.insert_category(sheet.title)

        category_attrs_list = []
        for column in range(6, sheet.max_column + 1):
            attr = sheet.cell(row=2, column=column).value
            self._check_attr(attr)
            category_attrs_list.append(self.prod_attr_ids[attr])

        if self.cli_args.bulk:
            self._write_bulk_products_on_db(sheet, category_id, category_attrs_list)
        else:
            self._write_products_on_db(sheet, category_id, category_attrs_list)

    def _write_products_on_db(self, sheet, category_id, category_attrs_list):
        with self.database as cursor:
            for row in range(3, sheet.max_row + 1):
                brand = sheet.cell(row=row, column=3).value
                self._check_brand(brand)
                product = {
                    'category_id': category_id,
                    'brand_id': self.brand_ids[brand],
                    'article': sheet.cell(row=row, column=2).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'price': sheet.cell(row=row, column=5).value,
                }
                attr_values = [sheet.cell(row=row, column=i).value for i in range(6, sheet.max_column + 1)]
                adv_attrs = {
                    attr_id: attr_value for (attr_id, attr_value) in zip(category_attrs_list, attr_values)
                }
                self.database.pure_insert_product(cursor, product, adv_attrs)

    def _write_bulk_products_on_db(self, sheet, category_id, category_attrs_list):
        with self.database as cursor:
            start = time()
            products = []
            advanced_attributes = []
            for row in range(3, sheet.max_row + 1):
                brand = sheet.cell(row=row, column=3).value
                self._check_brand(brand)
                product = {
                    'category_id': category_id,
                    'brand_id': self.brand_ids[brand],
                    'article': sheet.cell(row=row, column=2).value,
                    'name': sheet.cell(row=row, column=4).value,
                    'price': sheet.cell(row=row, column=5).value,
                }
                products.append(product)
                attr_values = [sheet.cell(row=row, column=i).value for i in range(6, sheet.max_column + 1)]
                for attr_id, attr_value in zip(category_attrs_list, attr_values):
                    advanced_attributes.append(
                        {
                            'article': product['article'],
                            'attribute_id': attr_id,
                            'attribute_value': attr_value
                        }
                    )

            print(f"READ products data DONE in {time()-start:.3f}sec")
            start = time()
            self.database.pure_insert_bulk_products(cursor, products, advanced_attributes)
            print(f'WRITE products data on SQL DB {time()-start:.3f}sec')

    def _check_attr(self, attr):
        if attr not in self.prod_attr_ids:
            self.prod_attr_ids[attr] = self.database.insert_product_attr(attr)

    def _check_brand(self, brand):
        if brand not in self.brand_ids:
            self.brand_ids[brand] = self.database.insert_brand(brand)
